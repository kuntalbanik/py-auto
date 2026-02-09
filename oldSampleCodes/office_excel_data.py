import openpyxl

# List of available excels
excels = ['ORDER_BOOKING_2012-13.xlsx', 'ORDER_BOOKING_2013-14.xlsx', 'ORDER_BOOKING_2014-15.xlsx', 'ORDER_BOOKING_2015-16.xlsx', 'ORDER_BOOKING_2016-17.xlsx', 'ORDER_BOOKING_2017-18.xlsx', 'ORDER_BOOKING_2018-19.xlsx', 'ORDER_BOOKING_2019-20.xlsx', 'ORDER_BOOKING_2020-21.xlsx', 'ORDER_BOOKING_2021-22.xlsx', 'ORDER_BOOKING_2022-23.xlsx', 'ORDER_BOOKING_2023-24.xlsx', 'ORDER_BOOKING_2024-25.xlsx']
# excels = ['ORDER_BOOKING_2012-13.xlsx']

sheets = ['TMX-WWS', 'SPARES-FEMS', 'TMX-ENV', 'TMX-PHD']

# location of excel file
temp_path = "o_excels"

# open the workbook
# wb_obj = openpyxl.load_workbook(path)

# print list all sheets available
# print(wb_obj.sheetnames)



# get all values by row for sheet "TMX-WWS"
def return_val(sheet_obj, row, column, wb_output):
    for i in range(1, row + 1):
        output_sheet = wb_output.active
        
        none_counter = 0
        cell_obj = sheet_obj.cell(row=i, column=1)
        cell_obj2 = sheet_obj.cell(row=i, column=column)
        if cell_obj.value != None or cell_obj2.value != None:
            # Save all to new excel file
            col_one = output_sheet.cell(row=i, column=1)
            col_two = output_sheet.cell(row=i, column=2)
            # col_one.value = cell_obj.value
            # col_two.value = cell_obj2.value
            # print(col_one.value, col_two.value)
            output_sheet.append((cell_obj.value, cell_obj2.value))

            # .cell(row=i, column=11).value = str(cell_obj.value).upper()
            # print(str(cell_obj.value).upper()+" - "+str(cell_obj2.value))
            # pass
        
        if cell_obj.value == None:
            none_counter += 1
            if none_counter > 8 :
                pass
    



def sheet_select(wb_output):
    for excel in excels:
        path = temp_path+"\\"+excel
        wb_obj = openpyxl.load_workbook(path)
        for sheet in sheets:
            wb_obj.active = wb_obj.sheetnames.index(sheet)
            sheet_obj = wb_obj.active
            row = sheet_obj.max_row
            # column = sheet_obj.max_column

            if sheet_obj.title == "TMX-WWS":
                column = 6
                return_val(sheet_obj, row, column, wb_output)
            
            if sheet_obj.title == "SPARES-FEMS":
                column = 11
                return_val(sheet_obj, row, column, wb_output)

            if sheet_obj.title == "TMX-ENV":
                column = 4
                return_val(sheet_obj, row, column, wb_output)

            if sheet_obj.title == "TMX-PHD":
                column = 8
                return_val(sheet_obj, row, column, wb_output)


if __name__ == '__main__':
    path = "o_excels\output.xlsx"
    wb_output = openpyxl.load_workbook(path)
    # output_sheet = wb_output.active
    
    sheet_select(wb_output)
    
    wb_output.save(path)




# Save all to new excel file
# wb_obj.save("Sample.xlsx")


    # print(cell_obj.value)

# cell_obj = sheet_obj.cell(row=1, column=1)

# print(cell_obj.value)







