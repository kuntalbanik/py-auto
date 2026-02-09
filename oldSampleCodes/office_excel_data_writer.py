import openpyxl

path_to_write = "o_excels\PRERANA_customer_name.xlsx"
path_to_read = "o_excels\GST_required.xlsx"


# open the workbook to read
wb_obj_r = openpyxl.load_workbook(path_to_read)

# open the workbook to write
wb_obj_w = openpyxl.load_workbook(path_to_write)

# get the active sheet object to read
sheet_obj_r = wb_obj_r.active

# get the active sheet object to write
sheet_obj_w = wb_obj_w.active

row_w = sheet_obj_w.max_row
row_r = sheet_obj_r.max_row
# column = sheet_obj_r.max_column

read_data = {}
for i in range(1, row_r + 1):
    # print(sheet_obj_r.cell(row=i, column=1).value)
    # read_data.insert(sheet_obj_r.cell(row=i, column=1).value, sheet_obj_r.cell(row=i, column=2).value)
    read_data[sheet_obj_r.cell(row=i, column=1).value] = sheet_obj_r.cell(row=i, column=2).value


# print(type(read_data))
# for ky,val in read_data.items():
#     print("Key : ",ky," : Value -> ",val)




# print("Rows : ", row)
# print("Column : ", column)

# get all values column numner 6
# print("\Value of sixth column")
for i in range(1, row_w + 1):
    cell_obj = sheet_obj_w.cell(row=i, column=11).value
    # print(cell_obj)
    if read_data.get(cell_obj) is not None:
        # print(cell_obj)
        # print(read_data[cell_obj])
        sheet_obj_w.cell(row=i, column=12).value = read_data[cell_obj]
        # print("Key exists in the dictionary.")
    else:
        print("Key does not exist in the dictionary.")
    # print(cell_obj.value)

    # if sheet_obj.cell(row=i, column=11).value == None:
    #     sheet_obj.cell(row=i, column=11).value = "Hello Excel"


# Save all to new excel file
wb_obj_w.save(path_to_write)


    # print(cell_obj.value)

# cell_obj = sheet_obj.cell(row=1, column=1)

# print(cell_obj.value)

